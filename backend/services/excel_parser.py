"""Excel parser for refinery signals."""

from datetime import date, datetime
from io import BytesIO
from typing import Optional
from pydantic import BaseModel
from openpyxl import load_workbook


class ParsedSignal(BaseModel):
    """A single signal parsed from the Excel file."""
    signal_id: str
    load_date: date
    refinery_tank_name: str  # Refinery tank name
    volume: float


class ExcelParseResult(BaseModel):
    """Result from parsing an Excel file."""
    signals: list[ParsedSignal]
    errors: list[str]


def parse_signals_excel(file_content: bytes) -> ExcelParseResult:
    """
    Parse refinery signals from an Excel file.

    Expected columns (flexible header matching):
    - Signal ID / ID / Signal
    - Load Date / Date / Scheduled Date
    - Refinery Tank / Source Tank / Tank
    - Volume / Quantity / Amount
    """
    signals: list[ParsedSignal] = []
    errors: list[str] = []

    try:
        wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            return ExcelParseResult(signals=[], errors=["No active worksheet found"])

        # Find header row and column indices
        header_row = None
        col_map: dict[str, int] = {}

        # Column name variations
        signal_id_names = ['signal id', 'signal_id', 'id', 'signal']
        date_names = ['load date', 'load_date', 'date', 'scheduled date', 'scheduled_date']
        tank_names = ['refinery tank', 'refinery_tank', 'source tank', 'source_tank', 'tank']
        volume_names = ['volume', 'quantity', 'amount', 'vol']

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_lower = str(cell).lower().strip()

                if cell_lower in signal_id_names:
                    col_map['signal_id'] = col_idx
                    header_row = row_idx
                elif cell_lower in date_names:
                    col_map['load_date'] = col_idx
                    header_row = row_idx
                elif cell_lower in tank_names:
                    col_map['refinery_tank_name'] = col_idx
                    header_row = row_idx
                elif cell_lower in volume_names:
                    col_map['volume'] = col_idx
                    header_row = row_idx

            if header_row is not None:
                break

        # Validate we found required columns
        required_cols = ['signal_id', 'load_date', 'refinery_tank_name', 'volume']
        missing = [col for col in required_cols if col not in col_map]
        if missing:
            return ExcelParseResult(
                signals=[],
                errors=[f"Missing required columns: {', '.join(missing)}"]
            )

        # Parse data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            # Skip empty rows
            if all(cell is None for cell in row):
                continue

            try:
                signal_id = row[col_map['signal_id']]
                load_date_raw = row[col_map['load_date']]
                refinery_tank_name = row[col_map['refinery_tank_name']]
                volume_raw = row[col_map['volume']]

                # Validate required fields
                if signal_id is None or refinery_tank_name is None or volume_raw is None:
                    errors.append(f"Row {row_idx}: Missing required field(s)")
                    continue

                # Parse date
                load_date: Optional[date] = None
                if isinstance(load_date_raw, datetime):
                    load_date = load_date_raw.date()
                elif isinstance(load_date_raw, date):
                    load_date = load_date_raw
                elif isinstance(load_date_raw, str):
                    # Try common date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                        try:
                            load_date = datetime.strptime(load_date_raw.strip(), fmt).date()
                            break
                        except ValueError:
                            continue

                if load_date is None:
                    errors.append(f"Row {row_idx}: Invalid date format '{load_date_raw}'")
                    continue

                # Parse volume
                try:
                    volume = float(volume_raw)
                    if volume <= 0:
                        errors.append(f"Row {row_idx}: Volume must be positive")
                        continue
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: Invalid volume '{volume_raw}'")
                    continue

                signals.append(ParsedSignal(
                    signal_id=str(signal_id).strip(),
                    load_date=load_date,
                    refinery_tank_name=str(refinery_tank_name).strip(),
                    volume=volume
                ))

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        wb.close()

    except Exception as e:
        errors.append(f"Failed to parse Excel file: {str(e)}")

    return ExcelParseResult(signals=signals, errors=errors)
